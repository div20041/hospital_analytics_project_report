"""
build_excel.py
Creates a professional Excel workbook with:
  Sheet 1 — Executive Summary (KPIs + charts data)
  Sheet 2 — Department Analysis (formatted table + formulas)
  Sheet 3 — Monthly Revenue (trend data)
  Sheet 4 — Patient Segments
  Sheet 5 — Doctor Scorecard
  Sheet 6 — Raw Data Dictionary
"""

import pandas as pd
import numpy as np
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import os, sys
sys.path.insert(0, "..")

# ── Load data ─────────────────────────────────────────────────────────────────
conn = mysql.connector.connect(
    host     = "localhost",
    user     = "root",
    password = "MYSQL",
    database = "hospital_analytics"
)
admissions  = pd.read_sql("SELECT * FROM admissions",  conn)
patients    = pd.read_sql("SELECT * FROM patients",    conn)
doctors     = pd.read_sql("SELECT * FROM doctors",     conn)
procedures  = pd.read_sql("SELECT * FROM procedures",  conn)
conn.close()

admissions["admission_date"] = pd.to_datetime(admissions["admission_date"])

# ── Style helpers ──────────────────────────────────────────────────────────────
DARK_BLUE   = "0D3B6E"
MID_BLUE    = "0E7490"
LIGHT_BLUE  = "DBEAFE"
GREEN       = "16A34A"
LIGHT_GREEN = "DCFCE7"
ORANGE      = "D97706"
LIGHT_ORG   = "FEF3C7"
RED         = "DC2626"
LIGHT_RED   = "FEE2E2"
WHITE       = "FFFFFF"
GREY_LIGHT  = "F1F5F9"
GREY_MID    = "94A3B8"
BLACK       = "000000"

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color=BLACK):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def style_header_row(ws, row, cols, bg=DARK_BLUE, fg=WHITE, size=10):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(bg)
        cell.font = header_font(size=size, color=fg)
        cell.alignment = center()
        cell.border = border_thin()

def style_data_row(ws, row, cols, bg=WHITE, alternate=GREY_LIGHT, row_num=0):
    bg_use = alternate if row_num % 2 == 1 else bg
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(bg_use)
        cell.font = body_font()
        cell.alignment = left_align()
        cell.border = border_thin()

wb = Workbook()

# ════════════════════════════════════════════════════════
#  SHEET 1 — EXECUTIVE SUMMARY
# ════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:H1")
ws1["A1"] = "🏥  HOSPITAL REVENUE & PATIENT ANALYTICS — EXECUTIVE SUMMARY"
ws1["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws1["A1"].fill = fill(DARK_BLUE)
ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 40

ws1.merge_cells("A2:H2")
ws1["A2"] = "Analysis Period: Jan 2021 – Dec 2024  |  Data: 5,000 Admissions | 3,000 Patients | 10 Departments"
ws1["A2"].font = Font(name="Arial", size=10, italic=True, color=GREY_MID)
ws1["A2"].fill = fill("E0F2FE")
ws1["A2"].alignment = center()
ws1.row_dimensions[2].height = 20

# KPI cards — row 4 onwards
ws1.row_dimensions[3].height = 12
ws1["A4"] = "KEY PERFORMANCE INDICATORS"
ws1["A4"].font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
ws1.row_dimensions[4].height = 22

kpis = [
    ("A5:B7", "Total Revenue",        f"₹{admissions['total_bill'].sum()/1e7:.2f} Crore",   MID_BLUE,   LIGHT_BLUE),
    ("C5:D7", "Total Admissions",     f"{len(admissions):,}",                                GREEN,      LIGHT_GREEN),
    ("E5:F7", "Unique Patients",      f"{admissions['patient_id'].nunique():,}",              DARK_BLUE,  LIGHT_BLUE),
    ("G5:H7", "Avg Bill / Admission", f"₹{admissions['total_bill'].mean():,.0f}",             ORANGE,     LIGHT_ORG),
    ("A8:B10","Avg Length of Stay",   f"{admissions['length_of_stay'].mean():.1f} days",     MID_BLUE,   LIGHT_BLUE),
    ("C8:D10","Readmission Rate",     f"{admissions['readmitted_30days'].mean()*100:.1f}%",   RED,        LIGHT_RED),
    ("E8:F10","Insurance Coverage",   f"{admissions['insurance_covered'].sum()/admissions['total_bill'].sum()*100:.1f}%", GREEN, LIGHT_GREEN),
    ("G8:H10","Recovery Rate",        f"{(admissions['outcome']=='Recovered').mean()*100:.1f}%", ORANGE, LIGHT_ORG),
]

for merge_range, label, value, accent, bg in kpis:
    start_cell = merge_range.split(":")[0]
    end_cell   = merge_range.split(":")[1]
    row_s = int(''.join(filter(str.isdigit, start_cell)))
    col_s = ''.join(filter(str.isalpha, start_cell))

    # Fill background for all cells in range
    col_e_letter = ''.join(filter(str.isalpha, end_cell))
    row_e = int(''.join(filter(str.isdigit, end_cell)))
    from openpyxl.utils import column_index_from_string, get_column_letter
    for rr in range(row_s, row_e+1):
        for cc in range(column_index_from_string(col_s), column_index_from_string(col_e_letter)+1):
            ws1.cell(row=rr, column=cc).fill = fill(bg)

    # Merge and write label
    ws1.merge_cells(f"{col_s}{row_s}:{col_e_letter}{row_s}")
    cell_label = ws1.cell(row=row_s, column=column_index_from_string(col_s))
    cell_label.value = label
    cell_label.font  = Font(name="Arial", size=9, color=accent, bold=True)
    cell_label.alignment = Alignment(horizontal="left", vertical="top")
    cell_label.border = Border(
        left=Side(style="medium", color=accent),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="medium", color=accent),
    )

    # Merge and write value
    val_row = row_s + 1
    ws1.merge_cells(f"{col_s}{val_row}:{col_e_letter}{row_e}")
    cell_val = ws1.cell(row=val_row, column=column_index_from_string(col_s))
    cell_val.value = value
    cell_val.font  = Font(name="Arial", size=14, bold=True, color=accent)
    cell_val.alignment = center()
    cell_val.border = Border(
        left=Side(style="medium", color=accent),
        right=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="medium", color=accent),
    )

    for r in range(row_s, row_e+1):
        ws1.row_dimensions[r].height = 18

# Department summary table
ws1.row_dimensions[11].height = 12
ws1["A12"] = "DEPARTMENT REVENUE SUMMARY"
ws1["A12"].font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
ws1.row_dimensions[12].height = 22

dept_summary = (admissions.groupby("department_name")
                .agg(admissions=("admission_id","count"),
                     revenue=("total_bill","sum"),
                     avg_bill=("total_bill","mean"),
                     avg_los=("length_of_stay","mean"),
                     readmission=("readmitted_30days","mean"))
                .reset_index()
                .sort_values("revenue", ascending=False))

headers13 = ["Department","Admissions","Total Revenue (₹)","Avg Bill (₹)","Avg LOS (days)","Readmission %","Revenue Share %"]
for j, h in enumerate(headers13, 1):
    ws1.cell(row=13, column=j).value = h
style_header_row(ws1, 13, len(headers13))

total_rev = dept_summary["revenue"].sum()
for i, (_, row) in enumerate(dept_summary.iterrows()):
    r = 14 + i
    vals = [
        row["department_name"],
        int(row["admissions"]),
        f"₹{row['revenue']:,.0f}",
        f"₹{row['avg_bill']:,.0f}",
        round(row["avg_los"], 1),
        f"{row['readmission']*100:.1f}%",
        f"=C{r}/C{r + len(dept_summary) - i - 1 + (i + len(dept_summary))}",  # placeholder
    ]
    # Simple revenue share
    share = row["revenue"] / total_rev * 100
    ws1.cell(row=r, column=1).value = row["department_name"]
    ws1.cell(row=r, column=2).value = int(row["admissions"])
    ws1.cell(row=r, column=3).value = round(row["revenue"], 0)
    ws1.cell(row=r, column=3).number_format = '₹#,##0'
    ws1.cell(row=r, column=4).value = round(row["avg_bill"], 0)
    ws1.cell(row=r, column=4).number_format = '₹#,##0'
    ws1.cell(row=r, column=5).value = round(row["avg_los"], 1)
    ws1.cell(row=r, column=6).value = round(row["readmission"]*100, 1)
    ws1.cell(row=r, column=7).value = f"=C{r}/SUM(C14:C{14+len(dept_summary)-1})"
    ws1.cell(row=r, column=7).number_format = '0.0%'
    style_data_row(ws1, r, 7, row_num=i)

# Totals row
tr = 14 + len(dept_summary)
ws1.cell(row=tr, column=1).value = "TOTAL"
ws1.cell(row=tr, column=2).value = f"=SUM(B14:B{tr-1})"
ws1.cell(row=tr, column=3).value = f"=SUM(C14:C{tr-1})"
ws1.cell(row=tr, column=3).number_format = '₹#,##0'
ws1.cell(row=tr, column=4).value = f"=AVERAGE(D14:D{tr-1})"
ws1.cell(row=tr, column=4).number_format = '₹#,##0'
ws1.cell(row=tr, column=5).value = f"=AVERAGE(E14:E{tr-1})"
style_header_row(ws1, tr, 7, bg=DARK_BLUE)

set_col_widths(ws1, {"A":22,"B":14,"C":20,"D":16,"E":16,"F":16,"G":16})

# ════════════════════════════════════════════════════════
#  SHEET 2 — MONTHLY REVENUE
# ════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly Revenue")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:G1")
ws2["A1"] = "Monthly Revenue Trend Analysis (2021–2024)"
ws2["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws2["A1"].fill = fill(MID_BLUE)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 35

monthly = (admissions.groupby(admissions["admission_date"].dt.to_period("M"))
           .agg(admissions=("admission_id","count"),
                revenue=("total_bill","sum"),
                avg_bill=("total_bill","mean"),
                readmissions=("readmitted_30days","sum"))
           .reset_index())
monthly["ym"] = monthly["admission_date"].astype(str)

headers2 = ["Month","Admissions","Revenue (₹)","Avg Bill (₹)","Readmissions","MoM Growth %","Cumulative Revenue (₹)"]
for j, h in enumerate(headers2, 1):
    ws2.cell(row=2, column=j).value = h
style_header_row(ws2, 2, len(headers2), bg=MID_BLUE)

for i, (_, row) in enumerate(monthly.iterrows()):
    r = 3 + i
    ws2.cell(row=r, column=1).value = row["ym"]
    ws2.cell(row=r, column=2).value = int(row["admissions"])
    ws2.cell(row=r, column=3).value = round(row["revenue"], 0)
    ws2.cell(row=r, column=3).number_format = '₹#,##0'
    ws2.cell(row=r, column=4).value = round(row["avg_bill"], 0)
    ws2.cell(row=r, column=4).number_format = '₹#,##0'
    ws2.cell(row=r, column=5).value = int(row["readmissions"])
    if r > 3:
        ws2.cell(row=r, column=6).value = f"=(C{r}-C{r-1})/C{r-1}"
        ws2.cell(row=r, column=6).number_format = '0.0%'
    else:
        ws2.cell(row=r, column=6).value = "-"
    ws2.cell(row=r, column=7).value = f"=SUM(C3:C{r})"
    ws2.cell(row=r, column=7).number_format = '₹#,##0'
    style_data_row(ws2, r, 7, row_num=i)

# Add line chart
last_row = 2 + len(monthly)
chart = LineChart()
chart.title = "Monthly Revenue Trend"
chart.style = 10
chart.y_axis.title = "Revenue (₹)"
chart.x_axis.title = "Month"
chart.height = 12; chart.width = 24
data_ref = Reference(ws2, min_col=3, min_row=2, max_row=last_row)
chart.add_data(data_ref, titles_from_data=True)
cats = Reference(ws2, min_col=1, min_row=3, max_row=last_row)
chart.set_categories(cats)
ws2.add_chart(chart, f"A{last_row + 3}")

set_col_widths(ws2, {"A":14,"B":14,"C":18,"D":16,"E":14,"F":14,"G":22})

# ════════════════════════════════════════════════════════
#  SHEET 3 — DOCTOR SCORECARD
# ════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Doctor Scorecard")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:H1")
ws3["A1"] = "Doctor Performance Scorecard"
ws3["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws3["A1"].fill = fill(DARK_BLUE)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 35

doc_perf = (admissions.merge(doctors[["doctor_id","doctor_name","specialization",
                                       "experience_years"]], on="doctor_id", how="left")
             .groupby(["doctor_id","doctor_name","specialization","experience_years"])
             .agg(patients=("admission_id","count"),
                  revenue=("total_bill","sum"),
                  avg_bill=("total_bill","mean"),
                  recoveries=("outcome", lambda x: (x=="Recovered").sum()),
                  readmissions=("readmitted_30days","sum"))
             .reset_index()
             .sort_values("revenue", ascending=False))
doc_perf["recovery_rate"] = doc_perf["recoveries"] / doc_perf["patients"]

headers3 = ["Doctor","Specialization","Experience (yrs)","Patients","Revenue (₹)",
            "Avg Bill (₹)","Recovery Rate %","Readmissions","Performance"]
for j, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=j).value = h
style_header_row(ws3, 2, len(headers3))

for i, (_, row) in enumerate(doc_perf.head(30).iterrows()):
    r = 3 + i
    perf = "⭐ Top" if row["revenue"] > doc_perf["revenue"].quantile(0.75) else "✓ Good" if row["revenue"] > doc_perf["revenue"].median() else "○ Avg"
    ws3.cell(row=r, column=1).value = row["doctor_name"]
    ws3.cell(row=r, column=2).value = row["specialization"]
    ws3.cell(row=r, column=3).value = int(row["experience_years"])
    ws3.cell(row=r, column=4).value = int(row["patients"])
    ws3.cell(row=r, column=5).value = round(row["revenue"], 0)
    ws3.cell(row=r, column=5).number_format = '₹#,##0'
    ws3.cell(row=r, column=6).value = round(row["avg_bill"], 0)
    ws3.cell(row=r, column=6).number_format = '₹#,##0'
    ws3.cell(row=r, column=7).value = round(row["recovery_rate"]*100, 1)
    ws3.cell(row=r, column=8).value = int(row["readmissions"])
    ws3.cell(row=r, column=9).value = perf
    style_data_row(ws3, r, 9, row_num=i)
    if perf == "⭐ Top":
        for col in range(1,10):
            ws3.cell(row=r, column=col).fill = fill(LIGHT_GREEN)

set_col_widths(ws3, {"A":22,"B":20,"C":16,"D":12,"E":18,"F":16,"G":16,"H":14,"I":12})

# ════════════════════════════════════════════════════════
#  SHEET 4 — DIAGNOSIS ANALYSIS
# ════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Diagnosis Analysis")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:G1")
ws4["A1"] = "Top Diagnoses — Frequency, Cost & Readmission Risk"
ws4["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws4["A1"].fill = fill(GREEN)
ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 35

diag = (admissions.groupby(["diagnosis","department_name"])
        .agg(cases=("admission_id","count"),
             avg_cost=("total_bill","mean"),
             total_rev=("total_bill","sum"),
             avg_los=("length_of_stay","mean"),
             readmission_rate=("readmitted_30days","mean"))
        .reset_index()
        .sort_values("cases", ascending=False)
        .head(25))

headers4 = ["Diagnosis","Department","Cases","Avg Cost (₹)","Total Revenue (₹)",
            "Avg LOS (days)","Readmission Rate %"]
for j, h in enumerate(headers4, 1):
    ws4.cell(row=2, column=j).value = h
style_header_row(ws4, 2, len(headers4), bg=GREEN)

for i, (_, row) in enumerate(diag.iterrows()):
    r = 3 + i
    ws4.cell(row=r, column=1).value = row["diagnosis"]
    ws4.cell(row=r, column=2).value = row["department_name"]
    ws4.cell(row=r, column=3).value = int(row["cases"])
    ws4.cell(row=r, column=4).value = round(row["avg_cost"], 0)
    ws4.cell(row=r, column=4).number_format = '₹#,##0'
    ws4.cell(row=r, column=5).value = round(row["total_rev"], 0)
    ws4.cell(row=r, column=5).number_format = '₹#,##0'
    ws4.cell(row=r, column=6).value = round(row["avg_los"], 1)
    ws4.cell(row=r, column=7).value = round(row["readmission_rate"]*100, 1)
    style_data_row(ws4, r, 7, row_num=i)
    # Highlight high readmission
    if row["readmission_rate"] > 0.15:
        ws4.cell(row=r, column=7).fill = fill(LIGHT_RED)
        ws4.cell(row=r, column=7).font = Font(name="Arial", size=10, bold=True, color=RED)

set_col_widths(ws4, {"A":26,"B":22,"C":10,"D":16,"E":20,"F":16,"G":18})

# ════════════════════════════════════════════════════════
#  SHEET 5 — DATA DICTIONARY
# ════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Data Dictionary")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:D1")
ws5["A1"] = "Data Dictionary — Hospital Analytics Dataset"
ws5["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws5["A1"].fill = fill(DARK_BLUE)
ws5["A1"].alignment = center()
ws5.row_dimensions[1].height = 30

fields = [
    ("admissions","admission_id","TEXT","Unique admission identifier (ADM00001)"),
    ("admissions","patient_id","TEXT","Patient identifier (PAT10000)"),
    ("admissions","doctor_id","TEXT","Treating doctor identifier"),
    ("admissions","department_name","TEXT","Hospital department name"),
    ("admissions","diagnosis","TEXT","Primary diagnosis"),
    ("admissions","admission_date","DATE","Date of admission (YYYY-MM-DD)"),
    ("admissions","discharge_date","DATE","Date of discharge"),
    ("admissions","length_of_stay","INTEGER","Days admitted"),
    ("admissions","severity","TEXT","Mild / Moderate / Severe / Critical"),
    ("admissions","total_bill","FLOAT","Total billing amount in ₹"),
    ("admissions","insurance_covered","FLOAT","Amount paid by insurance in ₹"),
    ("admissions","patient_paid","FLOAT","Amount paid by patient in ₹"),
    ("admissions","readmitted_30days","INTEGER","1 if readmitted within 30 days, else 0"),
    ("admissions","outcome","TEXT","Recovered / Improved / Referred / Expired"),
    ("patients","patient_id","TEXT","Unique patient identifier"),
    ("patients","age","INTEGER","Patient age in years"),
    ("patients","gender","TEXT","Male / Female"),
    ("patients","insurance","TEXT","Insurance provider name"),
    ("procedures","procedure_id","TEXT","Unique procedure identifier"),
    ("procedures","procedure_name","TEXT","Name of clinical procedure"),
    ("procedures","cost","FLOAT","Procedure cost in ₹"),
]

for j, h in enumerate(["Table","Column","Data Type","Description"], 1):
    ws5.cell(row=2, column=j).value = h
style_header_row(ws5, 2, 4)

for i, (tbl, col, dtype, desc) in enumerate(fields):
    r = 3 + i
    ws5.cell(row=r, column=1).value = tbl
    ws5.cell(row=r, column=2).value = col
    ws5.cell(row=r, column=3).value = dtype
    ws5.cell(row=r, column=4).value = desc
    style_data_row(ws5, r, 4, row_num=i)

set_col_widths(ws5, {"A":18,"B":22,"C":14,"D":55})

# ── Save ──────────────────────────────────────────────────────────────────────
os.makedirs("outputs", exist_ok=True)
out_path = "outputs/Hospital_Analytics_Report.xlsx"
wb.save(out_path)
print(f"✓ Excel workbook saved: {out_path}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")

# Recalculate formulas
import subprocess
result = subprocess.run(
    ["python", "/mnt/skills/public/xlsx/scripts/recalc.py", out_path],
    capture_output=True, text=True
)
print(f"  Recalc: {result.stdout.strip()}")
